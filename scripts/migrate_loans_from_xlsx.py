"""
Migrates loan data from the legacy workbook into the new database:

  1. Seeds `loan_types` from the ' Loan Types' sheet (5 real products:
     Capital, Short Term, Quick, Item, Christmas Items loans).
  2. Scans every 'annual_loan_N' table in the 'Annual Loan Database' sheet
     for historical disbursements and inserts them into `loans`.

As of this writing, the workbook's own 'Annual Loan List' sheet reports
"Total Loans: 0" -- there is genuinely no historical loan data to migrate
yet. Step 2 will therefore process 0 real rows today, but is written to
work correctly the moment real disbursement data exists in the sheet.

Usage:
    python migrate_loans_from_xlsx.py "/path/to/workbook.xlsx"

Requires DATABASE_URL to be set in the environment first, e.g.:
    export DATABASE_URL="postgresql://...neon connection string..."

Lessons baked in from the Members migration (see git history for the
original bugs this avoids):
  - Raw SQL is wrapped in sqlalchemy.text() -- SQLAlchemy 2.x raises
    ArgumentError on bare strings.
  - LoanStatus is stored by its .value ("active"), not its .name
    ("ACTIVE"), because models.py's LoanStatus column now uses
    values_callable. (MemberStatus/Gender still use .name -- see the
    NOTE in models.py if touching those instead.)
  - The file path argument is validated up front with a clear error,
    instead of letting openpyxl's low-level FileNotFoundError/
    InvalidFileException surface confusingly.
  - Inserts are idempotent: re-running this script does not create
    duplicate loan_types (matched by name) or duplicate loans (matched
    by member + loan type + disbursement date + principal).
"""

import os
import sys

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

# ---------------------------------------------------------------------------
# 0. Validate inputs up front, with clear errors, before touching Excel/DB
# ---------------------------------------------------------------------------

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise SystemExit(
        "DATABASE_URL is not set. Example:\n"
        '  export DATABASE_URL="postgresql://user:pass@host/db?sslmode=require"'
    )

if len(sys.argv) < 2:
    raise SystemExit(
        'Usage: python migrate_loans_from_xlsx.py "/path/to/workbook.xlsx"\n'
        "(Use the real path to your file -- don't leave this as a placeholder.)"
    )

WORKBOOK_PATH = sys.argv[1]

if not os.path.exists(WORKBOOK_PATH):
    raise SystemExit(
        f"File not found: {WORKBOOK_PATH!r}\n"
        "Double check the path. On Windows/Git Bash, drive letters are "
        "/c/, /d/, etc., and paths with spaces need to be quoted."
    )

engine = create_engine(DATABASE_URL)
Session = sessionmaker(bind=engine)


# ---------------------------------------------------------------------------
# 1. Seed loan_types from the ' Loan Types' sheet
# ---------------------------------------------------------------------------

def seed_loan_types(wb, session) -> dict:
    """
    Reads the loan_types table (A3:E8) from the ' Loan Types' sheet:
      A=S/NO, B=LOAN NAME, C=TENURE (MONTHS), D=FXD INTEREST, E=Charges

    Returns a dict mapping loan name -> id, for use when matching
    disbursement rows in step 2.
    """
    ws = wb[" Loan Types"]

    inserted, skipped = 0, 0
    name_to_id = {}

    for row in ws.iter_rows(min_row=4, max_row=8, values_only=False):
        name = row[1].value  # column B
        if not name or not str(name).strip():
            continue
        name = str(name).strip()

        tenure_months = row[2].value  # column C
        interest_rate = row[3].value  # column D, already a decimal fraction e.g. 0.1
        flat_charge = row[4].value or 0  # column E, e.g. 500 for Item Loan

        if tenure_months is None or interest_rate is None:
            skipped += 1
            continue

        existing = session.execute(
            text("SELECT id FROM loan_types WHERE name = :name"),
            {"name": name},
        ).fetchone()

        if existing:
            name_to_id[name] = existing[0]
            continue

        result = session.execute(
            text(
                """
                INSERT INTO loan_types
                    (id, name, interest_rate, tenure_months, flat_charge,
                     is_active, open_for_application, created_at, updated_at)
                VALUES
                    (gen_random_uuid(), :name, :interest_rate, :tenure_months,
                     :flat_charge, true, false, now(), now())
                RETURNING id
                """
            ),
            {
                "name": name,
                "interest_rate": interest_rate,
                "tenure_months": int(tenure_months),
                "flat_charge": flat_charge,
            },
        )
        new_id = result.fetchone()[0]

        # Also seed an initial rate version so the new effective-dated
        # rate history (see models.LoanTypeRateVersion) has a starting
        # point -- otherwise get_effective_terms() would silently fall
        # back to the loan_type's own cached fields, which works but
        # leaves no history to look back on.
        session.execute(
            text(
                """
                INSERT INTO loan_type_rate_versions
                    (id, loan_type_id, interest_rate, tenure_months, flat_charge, effective_from, created_at)
                VALUES
                    (gen_random_uuid(), :loan_type_id, :interest_rate, :tenure_months, :flat_charge, CURRENT_DATE, now())
                """
            ),
            {
                "loan_type_id": new_id,
                "interest_rate": interest_rate,
                "tenure_months": int(tenure_months),
                "flat_charge": flat_charge,
            },
        )

        name_to_id[name] = new_id
        inserted += 1

    session.commit()
    print(f"Loan types: inserted {inserted}, already existed {len(name_to_id) - inserted}, skipped {skipped}")
    return name_to_id


# ---------------------------------------------------------------------------
# 2. Scan every annual_loan_N table for historical disbursements
# ---------------------------------------------------------------------------

# Maps the Loan Disbursement sheet's column headers to the loan_types.name
# values seeded above. NOTE: 'CR  R&O' has a column in the disbursement
# sheet but no corresponding entry in the ' Loan Types' sheet -- this is
# an inconsistency in the original workbook. Since there's no data in that
# column today, we skip it with a warning rather than guessing a rate.
COLUMN_TO_LOAN_TYPE_NAME = {
    "Capital Loan": "Capital Loan",
    "Short Term Loan": "Short Term Loan",
    "Quick Loan": "Quick Loan",
    "ITEM LOAN": "Item Loan",
    "CR  R&O": None,  # no matching loan type defined in the workbook
}


def migrate_loans(wb, session, name_to_id: dict):
    ws = wb["Annual Loan Database"]

    inserted, skipped_blank, skipped_no_type = 0, 0, 0
    warned_missing_types = set()

    annual_loan_tables = [t for t in ws.tables.values() if t.name.startswith("annual_loan_")]

    for table in annual_loan_tables:
        # table.ref is like "B7:Q57"; header row is the first row of that
        # range, data starts the row after.
        start_cell, end_cell = table.ref.split(":")
        start_row = int("".join(filter(str.isdigit, start_cell)))
        end_row = int("".join(filter(str.isdigit, end_cell)))
        start_col = "".join(filter(str.isalpha, start_cell))

        header_row = start_row
        headers = {}
        col_idx = ws[f"{start_col}{header_row}"].column
        for offset in range(16):  # B..Q is 16 columns
            cell = ws.cell(row=header_row, column=col_idx + offset)
            if cell.value:
                headers[str(cell.value).strip()] = col_idx + offset

        psn_col = headers.get("PSN")
        name_col = headers.get("DEDUCTION NAME")
        date_col = headers.get("Date")
        if not psn_col or not name_col:
            continue  # unexpected table shape, skip defensively

        for r in range(header_row + 1, end_row + 1):
            psn = ws.cell(row=r, column=psn_col).value
            if not psn or psn == 0:
                skipped_blank += 1
                continue

            # IFS-equivalent: find whichever loan-type column has an amount
            principal, loan_type_name = None, None
            for column_header, mapped_name in COLUMN_TO_LOAN_TYPE_NAME.items():
                col = headers.get(column_header)
                if not col:
                    continue
                value = ws.cell(row=r, column=col).value
                if isinstance(value, (int, float)) and value not in (0, None):
                    principal = value
                    loan_type_name = mapped_name
                    if mapped_name is None:
                        warned_missing_types.add(column_header)
                    break

            if principal is None:
                skipped_blank += 1
                continue
            if loan_type_name is None:
                skipped_no_type += 1
                continue

            loan_type_id = name_to_id.get(loan_type_name)
            if not loan_type_id:
                skipped_no_type += 1
                continue

            disbursement_date = ws.cell(row=r, column=date_col).value if date_col else None
            if not disbursement_date:
                skipped_blank += 1
                continue

            member = session.execute(
                text("SELECT id FROM members WHERE psn = :psn"),
                {"psn": str(psn).strip()},
            ).fetchone()
            if not member:
                skipped_no_type += 1  # member not found -- reuse counter, printed separately below
                continue

            # Idempotency check: skip if this exact loan already exists
            existing = session.execute(
                text(
                    """
                    SELECT id FROM loans
                    WHERE member_id = :member_id
                      AND loan_type_id = :loan_type_id
                      AND disbursement_date = :disbursement_date
                      AND principal = :principal
                    """
                ),
                {
                    "member_id": member[0],
                    "loan_type_id": loan_type_id,
                    "disbursement_date": disbursement_date,
                    "principal": principal,
                },
            ).fetchone()
            if existing:
                continue

            loan_type_row = session.execute(
                text("SELECT interest_rate, tenure_months, flat_charge FROM loan_types WHERE id = :id"),
                {"id": loan_type_id},
            ).fetchone()
            rate, tenure_months, flat_charge = loan_type_row

            # Interest-at-source model: interest is deducted from what's
            # disbursed, not added on top of what's repaid. See
            # app/loan_calc.py for the canonical version of this logic --
            # kept in sync here manually since this script runs standalone.
            interest_amount = principal * float(rate)
            net_disbursed = principal - interest_amount
            total_repayable = principal + float(flat_charge)
            monthly_installment = total_repayable / tenure_months

            member_account_row = session.execute(
                text("SELECT account_number FROM members WHERE id = :id"),
                {"id": member[0]},
            ).fetchone()
            disbursement_account_number = member_account_row[0] if member_account_row else None

            session.execute(
                text(
                    """
                    INSERT INTO loans
                        (id, member_id, loan_type_id, principal, interest_amount,
                         net_disbursed, total_repayable, monthly_installment,
                         disbursement_date, expected_end_date,
                         disbursement_account_number, amount_repaid, status,
                         created_at, updated_at)
                    VALUES
                        (gen_random_uuid(), :member_id, :loan_type_id, :principal,
                         :interest_amount, :net_disbursed, :total_repayable, :monthly_installment,
                         :disbursement_date,
                         :disbursement_date + (:tenure_months || ' months')::interval,
                         :disbursement_account_number,
                         0, 'active', now(), now())
                    """
                ),
                {
                    "member_id": member[0],
                    "loan_type_id": loan_type_id,
                    "principal": principal,
                    "interest_amount": interest_amount,
                    "net_disbursed": net_disbursed,
                    "total_repayable": total_repayable,
                    "monthly_installment": monthly_installment,
                    "disbursement_date": disbursement_date,
                    "tenure_months": tenure_months,
                    "disbursement_account_number": disbursement_account_number,
                },
            )
            inserted += 1

    session.commit()

    print(f"Loans: inserted {inserted}, skipped blank/no-amount rows {skipped_blank}, "
          f"skipped (no matching type or member) {skipped_no_type}")
    if warned_missing_types:
        print(
            f"Note: columns {sorted(warned_missing_types)} had disbursement columns "
            "in the sheet but no matching entry in the ' Loan Types' sheet, so any "
            "amounts in them were skipped rather than guessed at."
        )


def main():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    session = Session()

    name_to_id = seed_loan_types(wb, session)
    migrate_loans(wb, session, name_to_id)

    session.close()
    print("Done.")


if __name__ == "__main__":
    main()
