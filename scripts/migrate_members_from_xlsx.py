"""
One-off migration: reads the original 'membersTable' sheet from the legacy
workbook and inserts each row into the new `members` Postgres table.

Usage:
    DATABASE_URL=postgresql://... python migrate_members_from_xlsx.py path/to/workbook.xlsx

Notes:
- Run this AFTER the backend has created the `members` table
  (starting the FastAPI app once with Base.metadata.create_all is enough).
- Rows with no NAME are treated as blank/template rows and skipped.
- STATUS: the spreadsheet uses 1 = financial member; anything else is
  treated as non-financial, matching the COUNTIF logic in frontEnd!O11.
"""

import sys
import os

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise SystemExit("Set DATABASE_URL before running this script.")

if len(sys.argv) < 2:
    raise SystemExit("Usage: python migrate_members_from_xlsx.py <workbook.xlsx>")

WORKBOOK_PATH = sys.argv[1]

engine = create_engine(DATABASE_URL)
Session = sessionmaker(bind=engine)


def gender_map(value):
    if not value:
        return None
    value = str(value).strip().lower()
    if value.startswith("m"):
        return "Male"
    if value.startswith("f"):
        return "Female"
    return "Other"


def main():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["membersTable"]

    session = Session()
    inserted, skipped = 0, 0

    # Data starts at row 4 (row 3 is the header row in the original sheet)
    for row in ws.iter_rows(min_row=4, values_only=False):
        name = row[1].value  # column B = NAME
        if not name or not str(name).strip():
            skipped += 1
            continue

        psn = row[2].value  # column C = PSN
        if psn is None:
            skipped += 1
            continue

        record = {
            "psn": str(psn).strip(),
            "name": str(name).strip(),
            "bank_name": row[3].value,        # D
            "account_number": str(row[4].value) if row[4].value else None,  # E
            "gender": gender_map(row[5].value),  # F
            "department": row[6].value,       # G
            "phone": str(row[7].value) if row[7].value else None,  # H
            "email": row[8].value,            # I
            "next_of_kin": row[9].value,       # J
            "next_of_kin_phone": str(row[10].value) if row[10].value else None,  # K
            "status": "financial" if row[11].value == 1 else "non_financial",  # L
        }

        session.execute(
            """
            INSERT INTO members
                (id, psn, name, bank_name, account_number, gender, department,
                 phone, email, next_of_kin, next_of_kin_phone, status,
                 created_at, updated_at)
            VALUES
                (gen_random_uuid(), :psn, :name, :bank_name, :account_number,
                 :gender, :department, :phone, :email, :next_of_kin,
                 :next_of_kin_phone, :status, now(), now())
            ON CONFLICT (psn) DO NOTHING
            """,
            record,
        )
        inserted += 1

    session.commit()
    session.close()
    print(f"Done. Rows processed: {inserted}, skipped blank rows: {skipped}")


if __name__ == "__main__":
    main()
